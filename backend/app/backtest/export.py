"""Export backtest results to CSV, JSON, Excel, and HTML."""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from typing import Any

from openpyxl import Workbook


def export_trades_csv(result: dict) -> str:
    trades = result.get("trades") or []
    buf = io.StringIO()
    fields = ["id", "entry_date", "exit_date", "entry_price", "exit_price", "pnl", "pnl_pct", "hold_days"]
    writer = csv.DictWriter(buf, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for t in trades:
        writer.writerow(t)
    return buf.getvalue()


def export_result_json(result: dict) -> str:
    slim = {k: v for k, v in result.items() if k not in ("ohlcv",)}
    if "ohlcv" in result:
        slim["ohlcv_bars"] = len(result.get("ohlcv") or [])
    return json.dumps(slim, indent=2, default=str)


def export_result_xlsx(result: dict) -> bytes:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    metrics = [
        ("Symbol", result.get("symbol", "")),
        ("Period", result.get("period", "")),
        ("Total Return %", result.get("total_return", 0)),
        ("Sharpe", result.get("sharpe", 0)),
        ("Sortino", result.get("sortino", 0)),
        ("Max Drawdown %", result.get("max_drawdown", 0)),
        ("Win Rate %", result.get("win_rate", 0)),
        ("Profit Factor", result.get("profit_factor", 0)),
        ("Total Trades", len(result.get("trades") or [])),
        ("Generated", datetime.utcnow().isoformat()),
    ]
    for row_idx, (label, val) in enumerate(metrics, start=1):
        ws_sum.cell(row=row_idx, column=1, value=label)
        ws_sum.cell(row=row_idx, column=2, value=val)

    ws_trades = wb.create_sheet("Trades")
    trades = result.get("trades") or []
    headers = ["id", "entry_date", "exit_date", "entry_price", "exit_price", "pnl", "pnl_pct"]
    ws_trades.append(headers)
    for t in trades:
        ws_trades.append([t.get(h) for h in headers])

    if result.get("optimization_grid"):
        ws_opt = wb.create_sheet("Optimization")
        ws_opt.append(["param_set", "total_return", "sharpe", "max_drawdown", "win_rate"])
        for row in result["optimization_grid"]:
            ws_opt.append([
                json.dumps(row.get("params", {})),
                row.get("total_return"),
                row.get("sharpe"),
                row.get("max_drawdown"),
                row.get("win_rate"),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_optimization_html(result: dict, title: str = "Optimization Report") -> str:
    rows = result.get("optimization_grid") or []
    best = result.get("best_params") or {}
    table_rows = ""
    for i, row in enumerate(rows[:100]):
        params = json.dumps(row.get("params", {}))
        table_rows += (
            f"<tr><td>{i + 1}</td><td><code>{params}</code></td>"
            f"<td>{row.get('total_return', 0):+.2f}%</td>"
            f"<td>{row.get('sharpe', 0):.2f}</td>"
            f"<td>{row.get('max_drawdown', 0):.2f}%</td></tr>"
        )
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>{title}</title>
<style>
body{{font-family:Roboto,sans-serif;max-width:960px;margin:32px auto;padding:0 20px;color:#111}}
h1{{font-size:1.5rem}} table{{width:100%;border-collapse:collapse;margin-top:16px;font-size:13px}}
th,td{{border-bottom:1px solid #ddd;padding:8px;text-align:left}} th{{background:#f5f5f5}}
.best{{background:#ecfdf5;padding:12px;border-radius:8px;margin:16px 0}}
</style></head><body>
<h1>{title}</h1>
<div class="best"><strong>Best parameters:</strong> <code>{json.dumps(best)}</code></div>
<table><thead><tr><th>#</th><th>Parameters</th><th>Return</th><th>Sharpe</th><th>Max DD</th></tr></thead>
<tbody>{table_rows or "<tr><td colspan='5'>No optimization rows</td></tr>"}</tbody></table>
</body></html>"""
