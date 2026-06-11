"""Generate Streak Indicators research workbook with prebuild suggestions."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

OPERATORS = (
    "Equal To | Crosses Above | Crosses Below | Higher Than | Lower Than | "
    "Higher Than Equal To | Lower Than Equal To | Up By % | Down By %"
)

INDICATORS = [
    "SMA", "RSI", "RSI Moving Average", "EMA", "WMA", "DEMA", "TEMA",
    "Moving Average", "ADX", "ADX Moving Average", "Supertrend", "Williams %R",
    "Lower Bollinger Bands", "Middle Bollinger Bands", "Upper Bollinger Bands",
    "Aroon Oscillator", "VWAP", "VWAP Moving Average", "CCI", "Stochastic",
    "Volume", "Vortex Indicator", "Money Flow Index", "MFI Moving Average",
    "Previous N", "Opening Range", "Nth Candle", "Pivot Points", "Median Price",
    "Median Price Moving Average", "ATR Trailing Stoploss", "Close", "High",
    "Open", "Low", "Alligator", "Donchian Channel", "Keltner Channel",
    "Volume Oscillator", "Choppiness Index", "Ichimoku Cloud", "ATR",
    "Awesome Oscillator", "Bollinger Bandwidth", "Chaikin Money Flow",
    "Chande Momentum Oscillator", "Elder Force Index", "Ultimate Oscillator",
    "MACD", "+DI", "-DI", "OBV", "Stochastic RSI", "Parabolic SAR",
    "True Range", "Bollinger %B", "Aroon Up", "Aroon Down",
    "Linear Regression Forecast", "Standard Deviation", "Trend Intensity Index",
    "Momentum Indicator", "Price Rate Of Change", "Fractal Chaos Bands",
    "Swing Index", "Twiggs Money Flow", "OBV Moving Average", "Ehler Fisher",
    "Coppock Curve", "Time Series Forecast", "ATR Bands",
]

CATEGORIES = ["Scalping", "Trend Following", "Long-term", "Trend Reversal"]

# (suggestion, tag) per indicator
INDICATOR_SUGGESTIONS: dict[str, tuple[list[tuple[str, str]], list[tuple[str, str]]]] = {
    "SMA": (
        [
            ("SMA(close, 9, 0) crosses below SMA(close, 21, 0)", "Scalping"),
            ("SMA(close, 5, 0) crosses below SMA(close, 20, 0)", "Scalping"),
            ("SMA(close, 21, 0) crosses above Close(0)", "Scalping"),
            ("SMA(close, 20, 0) crosses below SMA(close, 50, 0)", "Trend Following"),
            ("SMA(close, 50, 0) crosses below SMA(close, 100, 0)", "Trend Following"),
            ("SMA(close, 50, 0) crosses below SMA(close, 200, 0)", "Trend Following"),
            ("SMA(close, 100, 0) crosses below SMA(close, 200, 0)", "Long-term"),
            ("SMA(close, 9, 0) lower than SMA(close, 21, 0)", "Long-term"),
            ("SMA(close, 20, 0) lower than SMA(close, 50, 0)", "Trend Reversal"),
            ("SMA(close, 50, 0) lower than SMA(close, 200, 0)", "Trend Reversal"),
        ],
        [
            ("SMA(close, 9, 0) crosses above SMA(close, 21, 0)", "Scalping"),
            ("SMA(close, 5, 0) crosses above SMA(close, 20, 0)", "Scalping"),
            ("Close(0) crosses above SMA(close, 21, 0)", "Scalping"),
            ("SMA(close, 20, 0) crosses above SMA(close, 50, 0)", "Trend Following"),
            ("SMA(close, 50, 0) crosses above SMA(close, 100, 0)", "Trend Following"),
            ("SMA(close, 50, 0) crosses above SMA(close, 200, 0)", "Trend Following"),
            ("SMA(close, 100, 0) crosses above SMA(close, 200, 0)", "Long-term"),
            ("SMA(close, 9, 0) higher than SMA(close, 21, 0)", "Long-term"),
            ("SMA(close, 20, 0) higher than SMA(close, 50, 0)", "Trend Reversal"),
            ("SMA(close, 50, 0) higher than SMA(close, 200, 0)", "Trend Reversal"),
        ],
    ),
    "RSI": (
        [
            ("RSI(14,0) crosses below 50", "Oscillator"),
            ("RSI(14,0) lower than RSI(14,-1)", "Oscillator"),
            ("RSI(14,0) crosses below RSI MA(14,simple,20,0)", "Oscillator"),
            ("RSI(14,0) lower than RSI MA(14,simple,20,0)", "Oscillator"),
        ],
        [
            ("RSI(14,0) crosses above 30", "Oscillator"),
            ("RSI(14,0) crosses above 50", "Oscillator"),
            ("RSI(14,0) higher than RSI(14,-1)", "Oscillator"),
            ("RSI(14,0) crosses above RSI MA(14,simple,20,0)", "Oscillator"),
            ("RSI(14,0) higher than RSI MA(14,simple,20,0)", "Oscillator"),
        ],
    ),
    "RSI Moving Average": (
        [
            ("RSI MA(14,simple,20,0) crosses below RSI MA(14,simple,40,0)", "Oscillator"),
        ],
        [
            ("RSI MA(14,simple,20,0) crosses above RSI MA(14,simple,40,0)", "Oscillator"),
        ],
    ),
    "EMA": (
        [
            ("EMA(close,9,0) crosses below EMA(close,21,0)", "Short-Term Bearish Crossover"),
            ("EMA(close,5,0) crosses below EMA(close,20,0)", "Fast EMA Bearish Crossover"),
            ("EMA(close,21,0) crosses above Close(0)", "Price Breakdown"),
            ("EMA(close,20,0) crosses below EMA(close,50,0)", "Bearish Crossover"),
            ("EMA(close,50,0) crosses below EMA(close,100,0)", "Momentum Breakdown"),
            ("EMA(close,50,0) crosses below EMA(close,200,0)", "Death Cross"),
            ("EMA(close,100,0) crosses below EMA(close,200,0)", "Slow EMA Bearish Crossover"),
            ("EMA(close,9,0) lower than EMA(close,21,0)", "Short-Term Bearish Bias"),
            ("EMA(close,50,0) crosses above Close(0)", "Bearish Breakdown"),
            ("EMA(close,20,0) lower than EMA(close,50,0)", "Down Trend Pressure"),
            ("EMA(close,200,0) crosses above Close(0)", "Strong Bearish Bias"),
            ("EMA(close,50,0) lower than EMA(close,200,0)", "Bearish Alignment"),
        ],
        [
            ("EMA(close,9,0) crosses above EMA(close,21,0)", "Short-Term Bullish Crossover"),
            ("EMA(close,5,0) crosses above EMA(close,20,0)", "Fast EMA Bullish Crossover"),
            ("EMA(close,21,0) crosses below Close(0)", "Price Breakout"),
            ("EMA(close,20,0) crosses above EMA(close,50,0)", "Trend Strengthening"),
            ("EMA(close,50,0) crosses above EMA(close,100,0)", "Momentum Build-Up"),
            ("EMA(close,50,0) crosses above EMA(close,200,0)", "Golden Cross"),
            ("EMA(close,100,0) crosses above EMA(close,200,0)", "Trend Strengthening"),
            ("EMA(close,9,0) higher than EMA(close,21,0)", "Short-Term Uptrend Holding"),
            ("EMA(close,50,0) crosses below Close(0)", "Mid-Term Breakout"),
            ("EMA(close,20,0) higher than EMA(close,50,0)", "Trend Direction Up"),
            ("EMA(close,200,0) crosses below Close(0)", "Long-Term Breakout"),
            ("EMA(close,50,0) higher than EMA(close,200,0)", "Bullish Alignment"),
        ],
    ),
    "DEMA": (
        [
            ("DEMA(close,9,0) crosses below DEMA(close,21,0)", "Bearish Crossover"),
            ("DEMA(close,5,0) crosses below DEMA(close,20,0)", "Fast Momentum Drop"),
            ("DEMA(close,21,0) crosses above Close(0)", "Bearish Confirmation"),
            ("DEMA(close,20,0) crosses below DEMA(close,50,0)", "Bearish Trend Start"),
            ("DEMA(close,50,0) crosses below DEMA(close,100,0)", "Trend Reversal"),
            ("DEMA(close,50,0) crosses below DEMA(close,200,0)", "Death Cross"),
            ("DEMA(close,100,0) crosses below DEMA(close,200,0)", "Trend Weakness"),
            ("DEMA(close,9,0) lower than DEMA(close,21,0)", "Bearish Bias"),
            ("DEMA(close,50,0) crosses above Close(0)", "Bearish Confirmation"),
            ("DEMA(close,20,0) lower than DEMA(close,50,0)", "Bearish Momentum"),
            ("DEMA(close,200,0) crosses above Close(0)", "Bearish Momentum"),
            ("DEMA(close,50,0) lower than DEMA(close,200,0)", "Downside Momentum"),
        ],
        [
            ("DEMA(close,9,0) crosses above DEMA(close,21,0)", "Trend Reversal"),
            ("DEMA(close,5,0) crosses above DEMA(close,20,0)", "Fast Momentum"),
            ("DEMA(close,21,0) crosses below Close(0)", "Breakout Confirmation"),
            ("DEMA(close,20,0) crosses above DEMA(close,50,0)", "Bullish Trend Start"),
            ("DEMA(close,50,0) crosses above DEMA(close,100,0)", "Momentum Shift"),
            ("DEMA(close,50,0) crosses above DEMA(close,200,0)", "Golden Crossover"),
            ("DEMA(close,100,0) crosses above DEMA(close,200,0)", "Bullish Signal"),
            ("DEMA(close,9,0) higher than DEMA(close,21,0)", "Bullish Bias"),
            ("DEMA(close,50,0) crosses below Close(0)", "Bullish Confirmation"),
            ("DEMA(close,20,0) higher than DEMA(close,50,0)", "Uptrend Continuation"),
            ("DEMA(close,200,0) crosses below Close(0)", "Long-Term Bullish Confirmation"),
            ("DEMA(close,50,0) higher than DEMA(close,200,0)", "Long-Term Strength"),
        ],
    ),
    "Stochastic RSI": (
        [
            ("Stochastic RSI(14,14,Yes,0) crosses below 80", "Overbought"),
            ("Stochastic RSI(14,14,Yes,0) lower than 80", "Weakening Momentum"),
            ("Stochastic RSI(14,14,Yes,0) crosses below 50", "Momentum Loss"),
            ("Stochastic RSI(14,14,Yes,0) lower than 50", "Bearish Momentum"),
            (
                "Stochastic RSI(14,14,Yes,0) crosses below Stochastic RSI(14,14,Yes,0)",
                "Bearish Crossover",
            ),
            (
                "Stochastic RSI(14,14,Yes,0) lower than Stochastic RSI(14,14,Yes,0)",
                "Weak Momentum",
            ),
        ],
        [
            ("Stochastic RSI(14,14,Yes,0) crosses above 20", "Oversold Reversal"),
            ("Stochastic RSI(14,14,Yes,0) higher than 20", "Momentum Recovery"),
            ("Stochastic RSI(14,14,Yes,0) crosses above 50", "Momentum Strengthening"),
            ("Stochastic RSI(14,14,Yes,0) higher than 50", "Strong Momentum"),
            (
                "Stochastic RSI(14,14,Yes,0) crosses above Stochastic RSI(14,14,Yes,0)",
                "Momentum",
            ),
            (
                "Stochastic RSI(14,14,Yes,0) higher than Stochastic RSI(14,14,Yes,0)",
                "Trend Confirmation",
            ),
        ],
    ),
    "TEMA": (
        [
            ("TEMA(close,9,0) crosses below TEMA(close,21,0)", "Bearish Crossover"),
            ("TEMA(close,5,0) crosses below TEMA(close,20,0)", "Fast Momentum Drop"),
            ("TEMA(close,21,0) crosses above Close(0)", "Bearish Confirmation"),
            ("TEMA(close,20,0) crosses below TEMA(close,50,0)", "Bearish Trend Start"),
            ("TEMA(close,50,0) crosses below TEMA(close,100,0)", "Trend Reversal"),
            ("TEMA(close,50,0) crosses below TEMA(close,200,0)", "Death Cross"),
            ("TEMA(close,100,0) crosses below TEMA(close,200,0)", "Trend Weakness"),
            ("TEMA(close,9,0) lower than TEMA(close,21,0)", "Bearish Bias"),
            ("TEMA(close,50,0) crosses above Close(0)", "Bearish Confirmation"),
            ("TEMA(close,20,0) lower than TEMA(close,50,0)", "Bearish Momentum"),
            ("TEMA(close,200,0) crosses above Close(0)", "Bearish Momentum"),
            ("TEMA(close,50,0) lower than TEMA(close,200,0)", "Downside Momentum"),
        ],
        [
            ("TEMA(close,9,0) crosses above TEMA(close,21,0)", "Trend Reversal"),
            ("TEMA(close,5,0) crosses above TEMA(close,20,0)", "Fast Momentum"),
            ("TEMA(close,21,0) crosses below Close(0)", "Breakout Confirmation"),
            ("TEMA(close,20,0) crosses above TEMA(close,50,0)", "Bullish Trend Start"),
            ("TEMA(close,50,0) crosses above TEMA(close,100,0)", "Momentum Shift"),
            ("TEMA(close,50,0) crosses above TEMA(close,200,0)", "Golden Crossover"),
            ("TEMA(close,100,0) crosses above TEMA(close,200,0)", "Bullish Signal"),
            ("TEMA(close,9,0) higher than TEMA(close,21,0)", "Bullish Bias"),
            ("TEMA(close,50,0) crosses below Close(0)", "Bullish Confirmation"),
            ("TEMA(close,20,0) higher than TEMA(close,50,0)", "Uptrend Continuation"),
            ("TEMA(close,200,0) crosses below Close(0)", "Long-Term Bullish Confirmation"),
            ("TEMA(close,50,0) higher than TEMA(close,200,0)", "Long-Term Strength"),
        ],
    ),
    "Moving Average": (
        [
            (
                "Moving average(close,9,hull,0) crosses below Moving average(close,21,hull,0)",
                "Bearish Crossover",
            ),
            (
                "Moving average(close,5,hull,0) crosses below Moving average(close,20,hull,0)",
                "Fast Momentum Drop",
            ),
            ("Moving average(close,21,hull,0) crosses above Close(0)", "Bearish Confirmation"),
            (
                "Moving average(close,20,hull,0) crosses below Moving average(close,50,hull,0)",
                "Bearish Trend Start",
            ),
            (
                "Moving average(close,50,hull,0) crosses below Moving average(close,100,hull,0)",
                "Trend Reversal",
            ),
            (
                "Moving average(close,50,hull,0) crosses below Moving average(close,200,hull,0)",
                "Death Cross",
            ),
            (
                "Moving average(close,100,hull,0) crosses below Moving average(close,200,hull,0)",
                "Trend Weakness",
            ),
            (
                "Moving average(close,9,hull,0) lower than Moving average(close,21,hull,0)",
                "Bearish Bias",
            ),
            ("Moving average(close,50,hull,0) crosses above Close(0)", "Bearish Confirmation"),
            (
                "Moving average(close,20,hull,0) lower than Moving average(close,50,hull,0)",
                "Bearish Momentum",
            ),
            ("Moving average(close,200,hull,0) crosses above Close(0)", "Bearish Momentum"),
            (
                "Moving average(close,50,hull,0) lower than Moving average(close,200,hull,0)",
                "Downside Momentum",
            ),
        ],
        [
            (
                "Moving average(close,9,hull,0) crosses above Moving average(close,21,hull,0)",
                "Trend Reversal",
            ),
            (
                "Moving average(close,5,hull,0) crosses above Moving average(close,20,hull,0)",
                "Fast Momentum",
            ),
            ("Moving average(close,21,hull,0) crosses below Close(0)", "Breakout Confirmation"),
            (
                "Moving average(close,20,hull,0) crosses above Moving average(close,50,hull,0)",
                "Bullish Trend Start",
            ),
            (
                "Moving average(close,50,hull,0) crosses above Moving average(close,100,hull,0)",
                "Momentum Shift",
            ),
            (
                "Moving average(close,50,hull,0) crosses above Moving average(close,200,hull,0)",
                "Golden Crossover",
            ),
            (
                "Moving average(close,100,hull,0) crosses above Moving average(close,200,hull,0)",
                "Bullish Signal",
            ),
            (
                "Moving average(close,9,hull,0) higher than Moving average(close,21,hull,0)",
                "Bullish Bias",
            ),
            ("Moving average(close,50,hull,0) crosses below Close(0)", "Bullish Confirmation"),
            (
                "Moving average(close,20,hull,0) higher than Moving average(close,50,hull,0)",
                "Uptrend Continuation",
            ),
            (
                "Moving average(close,200,hull,0) crosses below Close(0)",
                "Long-Term Bullish Confirmation",
            ),
            (
                "Moving average(close,50,hull,0) higher than Moving average(close,200,hull,0)",
                "Long-Term Strength",
            ),
        ],
    ),
    "ADX": (
        [
            ("ADX(14,0) crosses below 25", "Weak Trend"),
            ("ADX(14,0) lower than 25", "Low Volatility"),
            (
                "ADX(14,0) crosses below Adx MA(14,adx,simple,20,0)",
                "Momentum Loss",
            ),
            (
                "ADX(14,0) lower than Adx MA(14,adx,simple,20,0)",
                "Trend Weakening",
            ),
        ],
        [
            ("ADX(14,0) crosses above 25", "Strong Trend"),
            ("ADX(14,0) higher than 25", "High Momentum"),
            (
                "ADX(14,0) crosses above Adx MA(14,adx,simple,20,0)",
                "Momentum",
            ),
            (
                "ADX(14,0) higher than Adx MA(14,adx,simple,20,0)",
                "Strength Confirmation",
            ),
        ],
    ),
    "ADX Moving Average": (
        [
            (
                "Adx MA(14,adx,simple,9,0) crosses below Adx MA(14,adx,simple,21,0)",
                "Strength Confirmation",
            ),
        ],
        [
            (
                "Adx MA(14,adx,simple,9,0) crosses above Adx MA(14,adx,simple,21,0)",
                "Strength Confirmation",
            ),
        ],
    ),
    "Supertrend": (
        [
            ("Supertrend(7,3,0) crosses above Close(0)", "Bearish Reversal"),
            ("Supertrend(7,3,0) higher than Close(0)", "Bearish Confirmation"),
        ],
        [
            ("Supertrend(7,3,0) crosses below Close(0)", "Bullish Reversal"),
            ("Supertrend(7,3,0) lower than Close(0)", "Uptrend Active"),
        ],
    ),
    "Williams %R": (
        [
            ("Willr(14,0) crosses below -20", "Overbought Reversal"),
            ("Willr(14,0) crosses below -50", "Momentum Breakdown"),
            ("Willr(14,0) lower than -20", "Overbought Zone"),
        ],
        [
            ("Willr(14,0) crosses above -80", "Oversold Reversal"),
            ("Willr(14,0) crosses above -50", "Momentum Strengthening"),
            ("Willr(14,0) higher than -80", "Out Of Oversold"),
        ],
    ),
    "Lower Bollinger Bands": (
        [],
        [
            ("LBB(Close,20,2,simple,0) crosses below Close(0)", "LBB Breakout"),
            ("LBB(Close,20,2,simple,0) higher than Close(0)", "Oversold Bounce"),
            ("LBB(Close,20,2,simple,0) lower than Close(0)", "Above LBB"),
            ("LBB(Close,20,2,simple,0) crosses above Close(0)", "Oversold Zone"),
        ],
    ),
    "Middle Bollinger Bands": (
        [
            ("MBB(Close,20,2,simple,0) crosses above Close(0)", "Bearish Momentum"),
            ("MBB(Close,20,2,simple,0) higher than Close(0)", "Trend Weakness"),
        ],
        [
            ("MBB(Close,20,2,simple,0) crosses below Close(0)", "Bullish Momentum"),
            ("MBB(Close,20,2,simple,0) lower than Close(0)", "Above Midband"),
        ],
    ),
    "Upper Bollinger Bands": (
        [
            ("UBB(Close,20,2,simple,0) lower than Close(0)", "Above UBB"),
            ("UBB(Close,20,2,simple,0) crosses above Close(0)", "Overbought Zone"),
        ],
        [
            ("UBB(Close,20,2,simple,0) crosses below Close(0)", "Reversal Signal"),
            ("UBB(Close,20,2,simple,0) higher than Close(0)", "Price Below UBB"),
        ],
    ),
    "Aroon Oscillator": (
        [
            ("Aroon oscillator(14,0) lower than 0", "Downtrend Strength"),
            ("Aroon oscillator(14,0) crosses below 0", "Bearish Reversal"),
        ],
        [
            ("Aroon oscillator(14,0) higher than 0", "Uptrend Strength"),
            ("Aroon oscillator(14,0) crosses above 0", "Bullish Reversal"),
        ],
    ),
    "VWAP": (
        [
            ("VWAP(0) crosses below Close(0)", "Breakdown"),
            ("VWAP(0) lower than Close(0)", "Trend Following"),
            ("VWAP(0) crosses below VWAP MA(simple,20,0)", "Bearish"),
        ],
        [
            ("VWAP(0) crosses above Close(0)", "Breakdown"),
            ("VWAP(0) higher than Close(0)", "Bearish"),
            ("VWAP(0) crosses above VWAP MA(simple,20,0)", "Breakout"),
        ],
    ),
    "VWAP Moving Average": (
        [
            ("VWAP MA(simple,9,0) crosses below VWAP MA(simple,21,0)", "Breakout"),
        ],
        [
            ("VWAP MA(simple,9,0) crosses above VWAP MA(simple,21,0)", "Breakout"),
        ],
    ),
    "CCI": (
        [
            ("CCI(20,0) crosses below 0", "Momentum Shift"),
            ("CCI(20,0) crosses below -100", "Overbought Exit"),
            ("CCI(20,0) lower than -100", "Momentum"),
        ],
        [
            ("CCI(20,0) crosses above 0", "Reversal"),
            ("CCI(20,0) crosses above 100", "Reversal"),
            ("CCI(20,0) higher than 100", "Momentum"),
        ],
    ),
    "Stochastic": (
        [
            (
                "Stochastic(14,3,Yes,0) crosses below Stochastic(14,3,Yes,0)",
                "Bearish Reversal",
            ),
            ("Stochastic(14,3,Yes,0) crosses below 80", "Overbought Reversal"),
            ("Stochastic(14,3,Yes,0) crosses below 50", "Momentum Weakening"),
            ("Stochastic(14,3,Yes,0) lower than 50", "Downtrend Momentum"),
            ("Stochastic(14,3,Yes,0) lower than 80", "Early Bearish Signal"),
        ],
        [
            (
                "Stochastic(14,3,Yes,0) crosses above Stochastic(14,3,Yes,0)",
                "Bullish Reversal",
            ),
            ("Stochastic(14,3,Yes,0) crosses above 20", "Oversold Reversal"),
            ("Stochastic(14,3,Yes,0) crosses above 50", "Momentum Strength"),
            ("Stochastic(14,3,Yes,0) higher than 20", "Bullish Bias"),
            ("Stochastic(14,3,Yes,0) higher than 50", "Uptrend Momentum"),
        ],
    ),
    "Volume": (
        [
            ("Volume(0) lower than 0", "Low Volume"),
            ("Volume(0) lower than Volume(-1)", "Falling Volume"),
            ("Volume(0) crosses below Moving average(volume,20,simple,0)", "Volume Drop"),
            (
                "Volume(0) lower than Moving average(volume,20,simple,0)",
                "Short-Term Volume Decline",
            ),
        ],
        [
            ("Volume(0) higher than 0", "High Volume"),
            ("Volume(0) crosses above Moving average(volume,20,simple,0)", "Volume Surge"),
            (
                "Volume(0) higher than Moving average(volume,20,simple,0)",
                "Short-Term Volume Spike",
            ),
            ("Volume(0) higher than Volume(-1)", "Rising Volume"),
        ],
    ),
    "OBV": (
        [
            ("OBV(0) lower than OBV(-1)", "Negative Directional Shift"),
        ],
        [
            ("OBV(0) higher than OBV(-1)", "Positive Directional Shift"),
        ],
    ),
    "Money Flow Index": (
        [
            ("MFI(20,0) crosses below 80", "Overbought Reversal"),
            ("MFI(20,0) crosses below 50", "Momentum Breakdown"),
            ("MFI(20,0) crosses below MFI MA(20,simple,20,0)", "Momentum Crossover Down"),
            ("MFI(20,0) lower than MFI MA(20,simple,20,0)", "Sustained Weakness"),
            ("MFI(20,0) lower than 80", "Exit Overbought"),
        ],
        [
            ("MFI(20,0) crosses above 20", "Oversold Reversal"),
            ("MFI(20,0) crosses above 50", "Midrange Breakout"),
            ("MFI(20,0) crosses above MFI MA(20,simple,20,0)", "Momentum Crossover"),
            ("MFI(20,0) higher than MFI MA(20,simple,20,0)", "Sustained Momentum"),
            ("MFI(20,0) higher than 20", "Exit Oversold"),
        ],
    ),
    "MFI Moving Average": (
        [
            ("MFI MA(20,simple,9,0) crosses below MFI MA(20,simple,21,0)", "Bearish Money Flow"),
        ],
        [
            ("MFI MA(20,simple,9,0) crosses above MFI MA(20,simple,21,0)", "Bullish Money Flow"),
        ],
    ),
    "Previous N": (
        [
            ("Prev N(Close,1,day) crosses above Close(0)", "Breakdown"),
            ("Prev N(Low,1,day) crosses above Close(0)", "Support Breakdown"),
            ("Prev N(Close,1,day) higher than Close(0)", "Weakness Confirmation"),
            ("Prev N(Low,1,day) higher than Close(0)", "Bearish Extension"),
            ("Prev N(Low,1,day) crosses above Low(0)", "New Low Signal"),
            ("Prev N(Low,1,day) higher than Low(0)", "Price Weakness"),
            ("Prev N(Volume,1,day) crosses above Volume(0)", "Volume Drop"),
            ("Prev N(Volume,1,day) higher than Volume(0)", "Lack Of Participation"),
        ],
        [
            ("Prev N(Close,1,day) crosses below Close(0)", "Breakout"),
            ("Prev N(High,1,day) crosses below Close(0)", "Resistance Breakout"),
            ("Prev N(Close,1,day) lower than Close(0)", "Price Strength"),
            ("Prev N(High,1,day) lower than Close(0)", "Bullish Extension"),
            ("Prev N(High,1,day) crosses below High(0)", "Intraday Breakout"),
            ("Prev N(High,1,day) lower than High(0)", "Trend Strength"),
            ("Prev N(Volume,1,day) crosses below Volume(0)", "Volume Spike"),
            ("Prev N(Volume,1,day) lower than Volume(0)", "Rising Interest"),
        ],
    ),
    "Opening Range": (
        [
            ("Opening Range(Low,15min) crosses above Close(0)", "Opening Range Breakdown"),
            ("Opening Range(Low,30min) crosses above Close(0)", "Opening Range Breakdown"),
            ("Opening Range(Low,15min) higher than Close(0)", "Breakdown Holding"),
            ("Opening Range(Low,30min) higher than Close(0)", "Breakdown Holding"),
            ("Opening Range(Volume,15min) crosses above Volume(0)", "Bearish Volume Spike"),
            ("Opening Range(Volume,30min) crosses above Volume(0)", "Bearish Volume Spike"),
        ],
        [
            ("Opening Range(High,15min) crosses below Close(0)", "Opening Range Breakout"),
            ("Opening Range(High,30min) crosses below Close(0)", "Opening Range Breakout"),
            ("Opening Range(High,15min) lower than Close(0)", "Breakout Holding"),
            ("Opening Range(High,30min) lower than Close(0)", "Breakout Holding"),
            ("Opening Range(Volume,15min) crosses below Volume(0)", "Volume Breakout"),
            ("Opening Range(Volume,30min) crosses below Volume(0)", "Volume Breakout"),
        ],
    ),
    "Nth Candle": (
        [
            ("Nth candle(High,15min,1) crosses above Close(0)", "Intraday Rejection"),
            ("Nth candle(High,30min,1) crosses above Close(0)", "Bearish Confirmation"),
            ("Nth candle(High,15min,1) higher than Close(0)", "Price Rejecting Resistance"),
            ("Nth candle(High,30min,1) higher than Close(0)", "Intraday Weakness"),
            (
                "Nth candle(High,15min,2) lower than Nth candle(High,15min,1)",
                "Lower High Formed",
            ),
        ],
        [
            ("Nth candle(High,15min,1) crosses below Close(0)", "Breakout Signal"),
            ("Nth candle(High,30min,1) crosses below Close(0)", "Momentum Spike"),
            (
                "Nth candle(High,15min,2) higher than Nth candle(High,15min,1)",
                "Momentum Build-up",
            ),
            ("Nth candle(High,15min,1) lower than Close(0)", "Short-Term Strength"),
            ("Nth candle(High,30min,1) lower than Close(0)", "Intraday Momentum"),
        ],
    ),
    "Pivot Points": (
        [
            ("Pivot point(pp,standard,No) crosses above Close(0)", "Pivot Breakdown"),
            ("Pivot point(1,standard,No) crosses above Close(0)", "S1 Breakdown"),
            ("Pivot point(2,standard,No) crosses above Close(0)", "S2 Breakdown"),
            ("Pivot point(3,standard,No) crosses above Close(0)", "S3 Breakdown"),
            ("Pivot point(pp,standard,No) higher than Close(0)", "Below Pivot"),
            ("Pivot point(1,standard,No) higher than Close(0)", "Below S1"),
            ("Pivot point(2,standard,No) higher than Close(0)", "Below S2"),
            ("Pivot point(3,standard,No) higher than Close(0)", "Oversold Breakdown"),
        ],
        [
            ("Pivot point(pp,standard,No) crosses below Close(0)", "Pivot Breakout"),
            ("Pivot point(1,standard,No) crosses below Close(0)", "R1 Breakout"),
            ("Pivot point(2,standard,No) crosses below Close(0)", "R2 Breakout"),
            ("Pivot point(3,standard,No) crosses below Close(0)", "R3 Breakout"),
            ("Pivot point(pp,standard,No) lower than Close(0)", "Holding Above Pivot"),
            ("Pivot point(1,standard,No) lower than Close(0)", "Above R1"),
            ("Pivot point(2,standard,No) lower than Close(0)", "Above R2"),
            ("Pivot point(3,standard,No) lower than Close(0)", "Extreme Strength"),
        ],
    ),
    "Vortex Indicator": (
        [
            ("Vortex(14,+vi,0) crosses below Vortex(14,-vi,0)", "Trend Reversal"),
            ("Vortex(14,+vi,0) crosses below 1", "Weakening Trend"),
            ("Vortex(14,+vi,0) lower than Vortex(14,-vi,0)", "Bearish Momentum"),
            ("Vortex(14,+vi,0) lower than 1", "Trend Weakness"),
        ],
        [
            ("Vortex(14,+vi,0) crosses above Vortex(14,-vi,0)", "Trend Reversal"),
            ("Vortex(14,+vi,0) crosses above 1", "Trend Strength"),
            ("Vortex(14,+vi,0) higher than Vortex(14,-vi,0)", "Trend Continuation"),
            ("Vortex(14,+vi,0) higher than 1", "Strong Uptrend"),
        ],
    ),
    "Median Price": (
        [
            ("Median Price(14,0) higher than Close(0)", "Short-Term Weakness"),
            (
                "Median Price MA(14,simple,20,0) crosses above Median Price(14,0)",
                "Bearish Reversal",
            ),
        ],
        [
            ("Median Price(14,0) lower than Close(0)", "Short-Term Strength"),
            (
                "Median Price MA(14,simple,20,0) crosses below Median Price(14,0)",
                "Bullish Reversal",
            ),
        ],
    ),
    "Median Price Moving Average": (
        [
            (
                "Median Price MA(14,simple,20,0) crosses above Median Price(14,0)",
                "Bearish Reversal",
            ),
        ],
        [
            (
                "Median Price MA(14,simple,20,0) crosses below Median Price(14,0)",
                "Bullish Reversal",
            ),
        ],
    ),
    "ATR Trailing Stoploss": (
        [
            ("ATR Trailing Stoploss(3,21,0) crosses above Close(0)", "Bearish Trend Shift"),
            ("ATR Trailing Stoploss(3,2,1,0) crosses above Close(0)", "ATR Trailing Stop Breakdown"),
        ],
        [
            ("ATR Trailing Stoploss(3,21,0) crosses below Close(0)", "Bullish Trend Shift"),
            ("ATR Trailing Stoploss(3,2,1,0) crosses below Close(0)", "ATR Trailing Stop Breakout"),
        ],
    ),
    "Close": (
        [
            ("Close(0) crosses below Close(-1)", "Bearish Reversal"),
            ("Close(0) lower than Close(-1)", "Bearish Close"),
            ("Close(0) lower than Open(0)", "Red Candle"),
            ("Close(0) crosses below DayHighLow(Low)", "Daily Low Breakdown"),
            ("Close(0) lower than DayHighLow(Low)", "Below Day Low"),
            ("Close(0) crosses below 0", "Level Breakdown"),
            ("Close(0) lower than 0", "Below Key Level"),
        ],
        [
            ("Close(0) crosses above Close(-1)", "Bullish Reversal"),
            ("Close(0) higher than Close(-1)", "Bullish Close"),
            ("Close(0) higher than Open(0)", "Green Candle"),
            ("Close(0) crosses above DayHighLow(High)", "Daily High Breakout"),
            ("Close(0) higher than DayHighLow(High)", "Above Day High"),
            ("Close(0) crosses above 0", "Level Breakout"),
            ("Close(0) higher than 0", "Above Key Level"),
        ],
    ),
    "High": (
        [],
        [
            ("High(0) crosses above High(-1)", "New High Formed"),
            ("High(0) higher than High(-1)", "Higher High"),
        ],
    ),
    "Open": (
        [
            ("Open(0) higher than Close(0)", "Red Candle"),
        ],
        [
            ("Open(0) lower than Close(0)", "Green Candle"),
        ],
    ),
    "Low": (
        [
            ("Low(0) crosses below Low(-1)", "New Low Formed"),
            ("Low(0) lower than Low(-1)", "Lower Low"),
        ],
        [
            ("Low(0) crosses above Low(-1)", "Early Strength"),
            ("Low(0) higher than Low(-1)", "Higher Low"),
        ],
    ),
    "Alligator": (
        [
            ("Alligator(lips) lower than Alligator(teeth)", "Downtrend Developing"),
            ("Alligator(teeth) lower than Alligator(jaw)", "Bearish Alignment"),
            ("Alligator(lips) lower than Alligator(jaw)", "Bearish Structure"),
            ("Alligator(lips) crosses below Alligator(teeth)", "Bearish Crossover"),
            ("Alligator(lips) crosses below Alligator(jaw)", "Breakdown"),
        ],
        [
            ("Alligator(lips) higher than Alligator(teeth)", "Uptrend Developing"),
            ("Alligator(teeth) higher than Alligator(jaw)", "Bullish Alignment"),
            ("Alligator(lips) higher than Alligator(jaw)", "Uptrend"),
            ("Alligator(lips) crosses above Alligator(teeth)", "Bullish Crossover"),
            ("Alligator(lips) crosses above Alligator(jaw)", "Breakout"),
        ],
    ),
    "Donchian Channel": (
        [
            ("Donchian Channel(Lower) crosses above Close(0)", "Breakdown"),
            ("Donchian Channel(Median) crosses above Close(0)", "Breakdown"),
            ("Donchian Channel(Lower) higher than Close(0)", "Breakdown"),
            ("Donchian Channel(Median) higher than Close(0)", "Breakdown"),
        ],
        [
            ("Donchian Channel(Upper) crosses below Close(0)", "Breakout"),
            ("Donchian Channel(Median) crosses below Close(0)", "Breakout"),
            ("Donchian Channel(Upper) lower than Close(0)", "Breakout"),
            ("Donchian Channel(Median) lower than Close(0)", "Breakout"),
        ],
    ),
    "Keltner Channel": (
        [
            ("Keltner(Bottom) crosses above Close(0)", "Bottom Band Breakdown"),
            ("Keltner(Median) crosses above Close(0)", "Mid Band Breakdown"),
            ("Keltner(Bottom) higher than Close(0)", "Keltner Breakdown"),
            ("Keltner(Median) higher than Close(0)", "Keltner Breakdown"),
        ],
        [
            ("Keltner(Top) crosses below Close(0)", "Upper Band Breakout"),
            ("Keltner(Median) crosses below Close(0)", "Midband Breakout"),
            ("Keltner(Top) lower than Close(0)", "Keltner Breakout"),
            ("Keltner(Median) lower than Close(0)", "Keltner Breakout"),
        ],
    ),
    "Volume Oscillator": (
        [
            ("Volume(0) lower than 0", "Low Volume"),
            ("Volume(0) lower than Volume(-1)", "Falling Volume"),
            ("Volume(0) crosses below Moving Average(volume,20)", "Volume Drop"),
            ("Volume(0) lower than Moving Average(volume,20)", "Short-Term Volume Decline"),
            ("Vortex(+VI) crosses below Vortex(-VI)", "Trend Reversal"),
            ("Vortex(+VI) crosses below 1", "Weakening Trend"),
            ("Vortex(+VI) lower than Vortex(-VI)", "Bearish Momentum"),
            ("Vortex(+VI) lower than 1", "Trend Weakness"),
            ("Pivot Point(PP) crosses above Close(0)", "Pivot Breakdown"),
            ("Pivot Point(S1) crosses above Close(0)", "S1 Breakdown"),
            ("Pivot Point(S2) crosses above Close(0)", "S2 Breakdown"),
            ("Pivot Point(S3) crosses above Close(0)", "S3 Breakdown"),
            ("Pivot Point(PP) higher than Close(0)", "Below Pivot"),
            ("Pivot Point(S1) higher than Close(0)", "Below S1"),
            ("Pivot Point(S2) higher than Close(0)", "Below S2"),
            ("Vortex Oscillator crosses below 0", "Falling Volume Momentum"),
            ("Vortex Oscillator lower than 0", "Falling Volume Momentum"),
        ],
        [
            ("Volume(0) higher than 0", "High Volume"),
            ("Volume(0) crosses above Moving Average(volume,20)", "Volume Surge"),
            ("Volume(0) higher than Moving Average(volume,20)", "Short-Term Volume Spike"),
            ("Volume(0) higher than Volume(-1)", "Rising Volume"),
            ("Vortex(+VI) crosses above Vortex(-VI)", "Trend Reversal"),
            ("Vortex(+VI) crosses above 1", "Trend Strength"),
            ("Vortex(+VI) higher than Vortex(-VI)", "Trend Continuation"),
            ("Vortex(+VI) higher than 1", "Strong Uptrend"),
            ("Pivot Point(PP) crosses below Close(0)", "Pivot Breakout"),
            ("Pivot Point(R1) crosses below Close(0)", "R1 Breakout"),
            ("Pivot Point(R2) crosses below Close(0)", "R2 Breakout"),
            ("Pivot Point(R3) crosses below Close(0)", "R3 Breakout"),
            ("Pivot Point(PP) lower than Close(0)", "Holding Above Pivot"),
            ("Pivot Point(R1) lower than Close(0)", "Above R1"),
            ("Pivot Point(R2) lower than Close(0)", "Above R2"),
            ("Vortex Oscillator crosses above 0", "Rising Volume Momentum"),
            ("Vortex Oscillator higher than 0", "Rising Volume Momentum"),
        ],
    ),
    "Choppiness Index": (
        [
            ("Choppiness Index(14) crosses above 62", "Consolidation"),
            ("Choppiness Index(14) higher than 62", "Sideways Market"),
        ],
        [
            ("Choppiness Index(14) crosses below 38", "Breakout Alert"),
            ("Choppiness Index(14) lower than 38", "Trending Market"),
        ],
    ),
    "Ichimoku Cloud": (
        [
            (
                "Ichimoku(9,26,52,26,conversion,yes,0) crosses above Close(0)",
                "Bearish Reversal",
            ),
            (
                "Ichimoku(9,26,52,26,conversion,yes,0) crosses below Ichimoku(9,26,52,26,base,yes,0)",
                "Bearish Crossover",
            ),
            (
                "Ichimoku(9,26,52,26,conversion,yes,0) lower than Ichimoku(9,26,52,26,base,yes,0)",
                "Trend Continuation",
            ),
            (
                "Ichimoku(9,26,52,26,leading span a,yes,0) lower than Ichimoku(9,26,52,26,leading span b,yes,0)",
                "Bearish Kumo Twist",
            ),
        ],
        [
            (
                "Ichimoku(9,26,52,26,conversion,yes,0) crosses below Close(0)",
                "Bullish Reversal",
            ),
            (
                "Ichimoku(9,26,52,26,conversion,yes,0) crosses above Ichimoku(9,26,52,26,base,yes,0)",
                "Bullish Crossover",
            ),
            (
                "Ichimoku(9,26,52,26,conversion,yes,0) higher than Ichimoku(9,26,52,26,base,yes,0)",
                "Trend Continuation",
            ),
            (
                "Ichimoku(9,26,52,26,leading span a,yes,0) higher than Ichimoku(9,26,52,26,leading span b,yes,0)",
                "Bullish Kumo Twist",
            ),
        ],
    ),
    "ATR": (
        [
            ("ATR(14,0) lower than 5", "Low Volatility"),
            ("ATR(14,0) crosses below 5", "Volatility Contraction"),
        ],
        [
            ("ATR(14,0) higher than 5", "High Volatility"),
            ("ATR(14,0) crosses above 5", "Volatility Expansion"),
        ],
    ),
    "ATR Bands": (
        [
            ("ATR Bands(5,2,5,Close,bottom,0) crosses below Close(0)", "ATR Bands Breakdown"),
            ("ATR Bands(5,2,5,Close,bottom,0) higher than Close(0)", "Below ATR Lower Band"),
            ("ATR Bands(5,2,5,Close,top,0) crosses above Close(0)", "ATR Bands Rejection"),
            ("ATR Bands(5,2,5,Close,top,0) higher than Close(0)", "Below ATR Upper Band"),
            ("ATR Bands(5,2.5,Close,Bottom,0) crosses below Close(0)", "ATR Bands"),
            ("ATR Bands(5,2.5,Close,Bottom,0) higher than Close(0)", "ATR Bands"),
            ("ATR Bands(5,2.5,Close,Top,0) crosses above Close(0)", "ATR Bands"),
            ("ATR Bands(5,2.5,Close,Top,0) higher than Close(0)", "ATR Bands"),
        ],
        [
            ("ATR Bands(5,2,5,Close,top,0) crosses below Close(0)", "ATR Bands Breakout"),
            ("ATR Bands(5,2,5,Close,top,0) lower than Close(0)", "Above ATR Upper Band"),
            ("ATR Bands(5,2,5,Close,bottom,0) lower than Close(0)", "Holding Above ATR Lower Band"),
            ("ATR Bands(5,2,5,Close,bottom,0) crosses above Close(0)", "ATR Bands Recovery"),
            ("ATR Bands(5,2.5,Close,Top,0) crosses below Close(0)", "ATR Bands"),
            ("ATR Bands(5,2.5,Close,Top,0) lower than Close(0)", "ATR Bands"),
            ("ATR Bands(5,2.5,Close,Bottom,0) lower than Close(0)", "ATR Bands"),
            ("ATR Bands(5,2.5,Close,Bottom,0) crosses above Close(0)", "ATR Bands"),
        ],
    ),
    "Awesome Oscillator": (
        [
            ("Awesome Oscillator(0) lower than 0", "Momentum"),
            ("Awesome Oscillator(0) crosses below 0", "Momentum"),
            ("Awesome Oscillator(0) lower than Awesome Oscillator(-3)", "Momentum"),
        ],
        [
            ("Awesome Oscillator(0) crosses above 0", "Momentum"),
            ("Awesome Oscillator(0) higher than Awesome Oscillator(-3)", "Momentum"),
        ],
    ),
    "Bollinger Bandwidth": (
        [
            ("Bollinger Bandwidth(Close,20,2,simple,0) lower than 5", "Volatility"),
            ("Bollinger Bandwidth(Close,20,2,simple,0) crosses below 5", "Volatility"),
        ],
        [
            ("Bollinger Bandwidth(Close,20,2,simple,0) higher than 5", "Volatility"),
            ("Bollinger Bandwidth(Close,20,2,simple,0) crosses above 5", "Volatility"),
        ],
    ),
    "Chaikin Money Flow": (
        [
            ("Chaikin MF(20,0) lower than 0", "Volume"),
            ("Chaikin MF(20,0) crosses below 0", "Volume"),
        ],
        [
            ("Chaikin MF(20,0) higher than 0", "Volume"),
            ("Chaikin MF(20,0) crosses above 0", "Volume"),
        ],
    ),
    "Chande Momentum Oscillator": (
        [
            ("CMO(9,0) lower than 0", "Momentum"),
            ("CMO(9,0) crosses below 0", "Momentum"),
            ("CMO(9,0) crosses below 50", "Momentum"),
        ],
        [
            ("CMO(9,0) higher than 0", "Momentum"),
            ("CMO(9,0) crosses above 0", "Momentum"),
            ("CMO(9,0) crosses above -50", "Momentum"),
        ],
    ),
    "Elder Force Index": (
        [
            ("EFI(13,0) lower than 0", "Volume Based"),
            ("EFI(13,0) crosses below 0", "Volume Based"),
            ("EFI(13,0) lower than EFI(13,-5)", "Volume Based"),
        ],
        [
            ("EFI(13,0) higher than 0", "Volume Based"),
            ("EFI(13,0) crosses above 0", "Volume Based"),
            ("EFI(13,0) higher than EFI(13,-5)", "Volume Based"),
        ],
    ),
    "Ultimate Oscillator": (
        [
            ("Ultimate Oscillator(7,14,28,0) crosses below 70", "Momentum"),
            ("Ultimate Oscillator(7,14,28,0) higher than 70", "Momentum"),
        ],
        [
            ("Ultimate Oscillator(7,14,28,0) crosses above 30", "Momentum"),
            ("Ultimate Oscillator(7,14,28,0) lower than 30", "Momentum"),
        ],
    ),
    "MACD": (
        [
            ("MACD(12,26,9,macd,0) crosses below MACD(12,26,9,signal,0)", "MACD Bearish Crossover"),
            ("MACD(12,26,9,histogram,0) crosses below 0", "MACD Histogram Negative"),
            ("MACD(12,26,9,macd,0) crosses below 0", "MACD Zero Line Breakdown"),
            ("MACD(12,26,9,macd,0) lower than MACD(12,26,9,signal,0)", "Negative Momentum"),
            ("MACD(12,26,9,histogram,0) lower than 0", "Sustained Bearish Momentum"),
        ],
        [
            ("MACD(12,26,9,macd,0) crosses above MACD(12,26,9,signal,0)", "MACD Bullish Crossover"),
            ("MACD(12,26,9,histogram,0) crosses above 0", "MACD Histogram Positive"),
            ("MACD(12,26,9,macd,0) crosses above 0", "Long-Term Bullish Signal"),
            ("MACD(12,26,9,histogram,0) higher than 0", "Sustained Bullish Momentum"),
            ("MACD(12,26,9,macd,0) higher than MACD(12,26,9,signal,0)", "Positive Momentum"),
        ],
    ),
    "+DI": (
        [
            ("Plus DI(14,0) crosses below Minus DI(14,0)", "Bearish Shift"),
            ("Plus DI(14,0) lower than Minus DI(14,0)", "Selling Pressure"),
        ],
        [
            ("Plus DI(14,0) crosses above Minus DI(14,0)", "Positive Directional Shift"),
            ("Plus DI(14,0) higher than Minus DI(14,0)", "Positive Momentum"),
        ],
    ),
    "-DI": (
        [
            ("Minus DI(14,0) crosses above Plus DI(14,0)", "Bearish Shift"),
            ("Minus DI(14,0) higher than Plus DI(14,0)", "Selling Pressure"),
        ],
        [
            ("Minus DI(14,0) crosses below Plus DI(14,0)", "Positive Directional Shift"),
            ("Minus DI(14,0) lower than Plus DI(14,0)", "Positive Directional Shift"),
        ],
    ),
    "Standard Deviation": (
        [
            ("StdDev(14,close,2,simple,0) lower than 0", "Momentum"),
            ("StdDev(14,close,2,simple,0) crosses below 0", "Momentum"),
        ],
        [
            ("StdDev(14,close,2,simple,0) higher than 0", "Momentum"),
            ("StdDev(14,close,2,simple,0) crosses above 0", "Momentum"),
        ],
    ),
    "Linear Regression Forecast": (
        [
            ("Lrf(close,14,0) crosses above Close(0)", "Trend Following"),
            ("Lrf(close,14,0) higher than Close(0)", "Trend Following"),
        ],
        [
            ("Lrf(close,14,0) crosses below Close(0)", "Trend Following"),
            ("Lrf(close,14,0) lower than Close(0)", "Trend Following"),
        ],
    ),
    "Aroon Down": (
        [
            ("Aroon Down(14,0) crosses above Aroon Up(14,0)", "Trend Following"),
            ("Aroon Down(14,0) higher than 70", "Trend Following"),
        ],
        [],
    ),
    "Aroon Up": (
        [],
        [
            ("Aroon Up(14,0) higher than 70", "Trend Following"),
            ("Aroon Up(14,0) crosses above Aroon Down(14,0)", "Trend Following"),
        ],
    ),
    "Bollinger %B": (
        [
            ("Bollinger %B(Close,20,2,simple,0) lower than 50", "Volatility"),
            ("Bollinger %B(Close,20,2,simple,0) crosses below 100", "Volatility"),
        ],
        [
            ("Bollinger %B(Close,20,2,simple,0) higher than 50", "Volatility"),
            ("Bollinger %B(Close,20,2,simple,0) crosses above 0", "Volatility"),
        ],
    ),
    "True Range": (
        [
            ("Trange(0) lower than 20", "Momentum"),
            ("Trange(0) crosses below 20", "Momentum"),
        ],
        [
            ("Trange(0) crosses above 20", "Momentum"),
            ("Trange(0) higher than 20", "Momentum"),
        ],
    ),
    "Parabolic SAR": (
        [
            ("Parabolic SAR(0.02,0.2,0) crosses above Close(0)", "Trend Following"),
            ("Parabolic SAR(0.02,0.2,0) higher than Close(0)", "Trend"),
        ],
        [
            ("Parabolic SAR(0.02,0.2,0) crosses below Close(0)", "Trend Following"),
            ("Parabolic SAR(0.02,0.2,0) lower than Close(0)", "Trend"),
        ],
    ),
    "Trend Intensity Index": (
        [
            ("TII(Close,TI,14,9,0) crosses below 80", "Overbought Reversal"),
            ("TII(Close,TI,14,9,0) crosses below 50", "Momentum Breakdown"),
            (
                "TII(Close,TI,14,9,0) crosses below TII(Close,Signal,14,9,0)",
                "Trend Reversal",
            ),
        ],
        [
            ("TII(Close,TI,14,9,0) crosses above 20", "Oversold Reversal"),
            (
                "TII(Close,TI,14,9,0) crosses above TII(Close,Signal,14,9,0)",
                "Momentum Reversal",
            ),
            ("TII(Close,TI,14,9,0) higher than 20", "Potential Reversal"),
            ("TII(Close,TI,14,9,0) crosses above 50", "Momentum Strengthening"),
            ("TII(Close,TI,14,9,0) lower than 80", "Out of Overbought"),
        ],
    ),
    "Momentum Indicator": (
        [
            ("Mom(14,0) crosses below 0", "Momentum Reversal"),
            ("Mom(14,0) lower than 0", "Momentum"),
        ],
        [
            ("Mom(14,0) crosses above 0", "Momentum"),
            ("Mom(14,0) higher than 0", "Momentum"),
        ],
    ),
    "Price Rate Of Change": (
        [
            ("Proc(close,14,0) crosses below 0", "Bearish Reversal"),
            ("Proc(close,14,0) lower than 0", "Negative Momentum"),
            ("Proc(close,14,0) lower than Proc(close,14,-1)", "Weakening Momentum"),
        ],
        [
            ("Proc(close,14,0) crosses above 0", "Bullish Reversal"),
            ("Proc(close,14,0) higher than 0", "Positive Momentum"),
            ("Proc(close,14,0) higher than Proc(close,14,-1)", "Increasing Momentum"),
        ],
    ),
    "Fractal Chaos Bands": (
        [
            ("Fractal Chaos Bands(Lower,0) higher than Close(0)", "Bearish Momentum"),
            (
                "Fractal Chaos Bands(Lower,0) lower than Fractal Chaos Bands(Lower,-1)",
                "Volatility Expansion",
            ),
        ],
        [
            ("Fractal Chaos Bands(Upper,0) lower than Close(0)", "Bullish Momentum"),
            ("Fractal Chaos Bands(Upper,0) crosses below Close(0)", "Trend Reversal"),
            (
                "Fractal Chaos Bands(Upper,0) higher than Fractal Chaos Bands(Upper,-1)",
                "Volatility Expansion",
            ),
            ("Fractal Chaos Bands(Lower,0) crosses above Close(0)", "Bullish Reversal"),
        ],
    ),
    "Swing Index": (
        [
            ("Swing(0,5,0) lower than 0", "Bearish Momentum"),
            ("Swing(0,5,0) crosses below 0", "Bearish Reversal"),
        ],
        [
            ("Swing(0,5,0) higher than 0", "Trend Strength"),
            ("Swing(0,5,0) crosses above 0", "Bullish Reversal"),
        ],
    ),
    "Twiggs Money Flow": (
        [
            ("Twiggs(21,0) lower than -0.2", "Bearish Distribution"),
            ("Twiggs(21,0) crosses below 0", "Bearish Reversal"),
        ],
        [
            ("Twiggs(21,0) higher than 0.2", "Bullish Accumulation"),
            ("Twiggs(21,0) crosses above 0", "Bullish Reversal"),
        ],
    ),
    "OBV Moving Average": (
        [
            ("OBV MA(Simple,20,0) crosses above OBV(0)", "Bearish Volume Crossover"),
            ("OBV MA(Exponential,20,0) crosses above OBV(0)", "Bearish Volume Momentum"),
            ("OBV MA(Simple,20,0) lower than OBV MA(Simple,20,-1)", "Falling Volume Trend"),
            (
                "OBV MA(Exponential,20,0) lower than OBV MA(Exponential,20,-1)",
                "Accelerating Distribution",
            ),
        ],
        [
            ("OBV MA(Simple,20,0) crosses below OBV(0)", "Bullish Volume Crossover"),
            ("OBV MA(Exponential,20,0) crosses below OBV(0)", "Bullish Volume Momentum"),
            ("OBV MA(Simple,20,0) higher than OBV MA(Simple,20,-1)", "Rising Volume Trend"),
            (
                "OBV MA(Exponential,20,0) higher than OBV MA(Exponential,20,-1)",
                "Accelerating Volume",
            ),
        ],
    ),
    "Ehler Fisher": (
        [
            (
                "Ehler Fisher(10,0,0) crosses below Ehler Fisher(10,Trigger,0)",
                "Bearish Momentum",
            ),
            ("Ehler Fisher(10,0,0) crosses below 0", "Momentum Breakdown"),
        ],
        [
            (
                "Ehler Fisher(10,0,0) crosses above Ehler Fisher(10,Trigger,0)",
                "Bullish Momentum",
            ),
            ("Ehler Fisher(10,0,0) crosses above 0", "Momentum Reversal"),
        ],
    ),
    "Coppock Curve": (
        [
            ("Coppock Curve(Close,10,11,14,0) crosses below 0", "Bearish Reversal"),
            (
                "Coppock Curve(Close,10,11,14,0) lower than Coppock Curve(Close,10,11,14,-1)",
                "Momentum Deterioration",
            ),
        ],
        [
            ("Coppock Curve(Close,10,11,14,0) crosses above 0", "Bullish Reversal"),
            (
                "Coppock Curve(Close,10,11,14,0) higher than Coppock Curve(Close,10,11,14,-1)",
                "Momentum Strengthening",
            ),
        ],
    ),
    "Time Series Forecast": (
        [
            ("TSF(14,Close,0) crosses above Close(0)", "Forecast"),
            ("TSF(14,Close,0) higher than Close(0)", "Forecast"),
        ],
        [
            ("TSF(14,Close,0) crosses below Close(0)", "Forecast"),
            ("TSF(14,Close,0) lower than Close(0)", "Forecast"),
        ],
    ),
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")


def _safe_range_name(indicator: str, direction: str) -> str:
    slug = indicator.replace(" ", "_").replace("%", "pct").replace("+", "plus").replace("-", "minus")
    return f"{direction}_{slug}"


def style_header_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_workbook() -> Workbook:
    wb = Workbook()
    lists = wb.active
    lists.title = "Lists"

    lists["A1"] = "Indicator"
    lists["B1"] = "Bearish Suggestion"
    lists["C1"] = "Tag"
    lists["D1"] = "Bullish Suggestion"
    lists["E1"] = "Tag"
    style_header_row(lists, 1, 5)

    # Master lookup table for VLOOKUP (all suggestions, unique per row)
    lists["G1"] = "Suggestion"
    lists["H1"] = "Tag"
    style_header_row(lists, 1, 8)

    lookup_entries: list[tuple[str, str]] = []
    list_row = 2

    for indicator, (bearish, bullish) in INDICATOR_SUGGESTIONS.items():
        start = list_row
        max_len = max(len(bearish), len(bullish))
        bearish_end = start + len(bearish) - 1
        bullish_end = start + len(bullish) - 1

        for i in range(max_len):
            lists.cell(list_row, 1, indicator if i == 0 else "")
            if i < len(bearish):
                suggestion, tag = bearish[i]
                lists.cell(list_row, 2, suggestion)
                lists.cell(list_row, 3, tag)
                lookup_entries.append((suggestion, tag))
            if i < len(bullish):
                suggestion, tag = bullish[i]
                lists.cell(list_row, 4, suggestion)
                lists.cell(list_row, 5, tag)
                lookup_entries.append((suggestion, tag))
            list_row += 1

        if max_len:
            list_row += 1  # blank separator row

        bearish_name = _safe_range_name(indicator, "Bearish")
        bullish_name = _safe_range_name(indicator, "Bullish")
        if bearish:
            wb.defined_names.add(
                DefinedName(bearish_name, attr_text=f"Lists!$B${start}:$B${bearish_end}")
            )
        if bullish:
            wb.defined_names.add(
                DefinedName(bullish_name, attr_text=f"Lists!$D${start}:$D${bullish_end}")
            )

    for i, (suggestion, tag) in enumerate(lookup_entries, start=2):
        lists.cell(i, 7, suggestion)
        lists.cell(i, 8, tag)
    lookup_end = 1 + len(lookup_entries)

    lists["J1"] = "Categories"
    style_header_row(lists, 1, 10)
    for i, cat in enumerate(CATEGORIES, start=2):
        lists.cell(i, 10, cat)

    lists.column_dimensions["A"].width = 22
    lists.column_dimensions["B"].width = 58
    lists.column_dimensions["C"].width = 28
    lists.column_dimensions["D"].width = 58
    lists.column_dimensions["E"].width = 28
    lists.column_dimensions["G"].width = 58
    lists.column_dimensions["H"].width = 28

    # --- Indicators sheet ---
    ws = wb.create_sheet("Indicators")
    headers = [
        "Indicator Name",
        "Supported Condition Operators",
        "Bearish Suggestion",
        "Tag",
        "Bias",
        "Bullish Suggestion",
        "Tag",
        "Bias",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    style_header_row(ws, 1, len(headers))

    indicator_rows: dict[str, int] = {}
    for row_idx, name in enumerate(INDICATORS, start=2):
        ws.cell(row_idx, 1, name)
        ws.cell(row_idx, 2, OPERATORS)
        indicator_rows[name] = row_idx

    for indicator in INDICATOR_SUGGESTIONS:
        row = indicator_rows[indicator]
        bearish_name = _safe_range_name(indicator, "Bearish")
        bullish_name = _safe_range_name(indicator, "Bullish")

        ws.cell(row, 4, f'=IF(C{row}="","",VLOOKUP(C{row},Lists!$G$2:$H${lookup_end},2,FALSE))')
        ws.cell(row, 5, f'=IF(C{row}="","","Bearish")')
        ws.cell(row, 7, f'=IF(F{row}="","",VLOOKUP(F{row},Lists!$G$2:$H${lookup_end},2,FALSE))')
        ws.cell(row, 8, f'=IF(F{row}="","","Bullish")')

        bearish_list, bullish_list = INDICATOR_SUGGESTIONS[indicator]
        if bearish_list:
            dv_bearish = DataValidation(
                type="list",
                formula1=f"={bearish_name}",
                allow_blank=True,
                showDropDown=False,
            )
            dv_bearish.error = f"Pick a bearish {indicator} suggestion from the list."
            dv_bearish.errorTitle = "Invalid suggestion"
            ws.add_data_validation(dv_bearish)
            dv_bearish.add(ws[f"C{row}"])

        if bullish_list:
            dv_bullish = DataValidation(
                type="list",
                formula1=f"={bullish_name}",
                allow_blank=True,
                showDropDown=False,
            )
            dv_bullish.error = f"Pick a bullish {indicator} suggestion from the list."
            dv_bullish.errorTitle = "Invalid suggestion"
            ws.add_data_validation(dv_bullish)
            dv_bullish.add(ws[f"F{row}"])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 58
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 12

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    return wb


if __name__ == "__main__":
    output = "Streak_Indicators_Final_Bullish_Bearish.xlsx"
    build_workbook().save(output)
    print(f"Saved: {output}")
